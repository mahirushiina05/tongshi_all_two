"""Question routes"""
from __future__ import annotations

from io import BytesIO

from typing import Optional

from fastapi import APIRouter, Depends, UploadFile, File, Query
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.security import get_current_user, require_role
from app.core.response import success
from app.core.exceptions import BusinessException
from app.core.upload_validation import validate_upload, ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE
from app.schemas.common import AuthUser, QuestionCreate, QuestionUpdate, CourseCreateRequest, CourseUpdateRequest
from app.services.question_service import (
    list_questions, create_question, update_question, delete_question,
    get_course_questions, create_course, add_public_course, update_course, delete_course,
    get_course_detail, import_questions_from_excel,
)
from app.services.course_response_service import build_course_list, build_course_detail

router = APIRouter(prefix="/questions", tags=["questions"])


def _format_question(q):
    return {
        "id": q.id,
        "type": q.type,
        "course_id": q.course_id,
        "course_name": q.course.name if q.course else "",
        "stem": q.stem,
        "options": q.options or [],
        "answer": q.answer,
        "explanation": q.explanation or "",
        "source_question_id": q.source_question_id,
        "is_synced": bool(q.source_question_id),
    }


@router.get("", summary="题目列表", description="按课程和题型筛选题目")
def get_questions(
    type: Optional[str] = None,
    course_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(get_current_user),
):
    teacher_id = current_user.id if current_user.role == "teacher" else None
    questions = list_questions(db, course_id, type, teacher_id)
    return success([_format_question(q) for q in questions])


@router.get("/course/{course_id}", summary="课程题目", description="获取指定课程的所有题目（用于测验）")
def get_course_questions_for_quiz(course_id: int, db: Session = Depends(get_db), _: AuthUser = Depends(get_current_user)):
    questions = get_course_questions(db, course_id)
    return success([_format_question(q) for q in questions])


@router.post("", summary="新增题目", description="教师端：创建选择题或填空题")
def add_question(data: QuestionCreate, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    q = create_question(db, data.model_dump(), current_user.id)
    return success({"id": q.id})


@router.put("/{question_id}", summary="编辑题目", description="教师端：修改指定题目的内容")
def edit_question(question_id: int, data: QuestionUpdate, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    q = update_question(db, question_id, data.model_dump(exclude_unset=True), current_user.id)
    if not q:
        raise BusinessException(404, "题目不存在")
    return success()


@router.delete("/{question_id}", summary="删除题目", description="教师端：删除指定题目")
def remove_question(question_id: int, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    if not delete_question(db, question_id, current_user.id):
        raise BusinessException(404, "题目不存在")
    return success()


@router.get("/courses", summary="课程列表", description="获取所有课程")
def get_courses(db: Session = Depends(get_db), current_user: AuthUser = Depends(get_current_user)):
    result = build_course_list(db, current_user)
    return success(result)


@router.post("/courses", summary="创建课程", description="教师端：创建新课程")
def add_course(data: CourseCreateRequest, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    course = create_course(db, data.name.strip(), current_user.id, False)
    return success({"id": course.id})


@router.post("/courses/{course_id}/add", summary="添加公共课程", description="教师端：将公共课程添加为自己的课程")
def add_public_course_to_teacher(course_id: int, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    course = add_public_course(db, course_id, current_user.id)
    return success({"id": course.id})


@router.get("/courses/{course_id}", summary="课程详情", description="返回课程信息和资料、题目、班级统计")
def get_course(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(get_current_user),
):
    teacher_id = current_user.id if current_user.role == "teacher" else None
    detail = get_course_detail(db, course_id, teacher_id)
    if not detail:
        raise BusinessException(404, "课程不存在")
    return success(build_course_detail(db, detail, current_user))


@router.put("/courses/{course_id}", summary="修改课程名称", description="教师端：修改指定课程的名称")
def edit_course(course_id: int, data: CourseUpdateRequest, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    course = update_course(db, course_id, data.name.strip(), current_user.id)
    if not course:
        raise BusinessException(404, "课程不存在")
    return success()


@router.delete("/courses/{course_id}", summary="删除课程", description="教师端：删除指定课程")
def remove_course(course_id: int, db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    if not delete_course(db, course_id, current_user.id):
        raise BusinessException(404, "课程不存在")
    return success()


def _build_question_template(question_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"
    ws.append(["题型", "课程名称", "题干", "选项（选择题用 | 分隔）", "答案", "解析"])
    if question_type == "choice":
        ws.append(["choice", "示例课程", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
    elif question_type == "fill":
        ws.append(["fill", "示例课程", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    else:
        ws.append(["choice", "示例课程", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
        ws.append(["fill", "示例课程", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _download_template_response(template_type: str):
    content = _build_question_template(template_type)
    filename_map = {
        "choice": "choice-question-template.xlsx",
        "fill": "fill-question-template.xlsx",
        "all": "question-template.xlsx",
    }
    filename = filename_map[template_type]
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/import/template", summary="下载题目导入模板", description="教师端：下载 Excel 批量导入模板（中文表头，支持选择题和填空题）")
def download_question_template(template_type: str = Query("all", pattern="^(all|choice|fill)$"), current_user: AuthUser = Depends(require_role("teacher"))):
    return _download_template_response(template_type)


@router.get("/import/template/choice", summary="下载选择题导入模板", description="教师端：下载选择题 Excel 模板")
def download_choice_question_template(current_user: AuthUser = Depends(require_role("teacher"))):
    return _download_template_response("choice")


@router.get("/import/template/fill", summary="下载填空题导入模板", description="教师端：下载填空题 Excel 模板")
def download_fill_question_template(current_user: AuthUser = Depends(require_role("teacher"))):
    return _download_template_response("fill")


@router.post("/import", summary="Excel 批量导入题目", description="教师端：上传 Excel 批量导入题目（.xlsx，表头：题型/课程名称/题干/选项/答案/解析）")
def import_questions(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: AuthUser = Depends(require_role("teacher"))):
    content = file.file.read()
    err = validate_upload(file.filename, len(
        content), allowed_extensions=ALLOWED_EXCEL_EXTENSIONS, max_size=MAX_EXCEL_SIZE)
    if err:
        raise BusinessException(400, err)
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(
        ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] if i < len(
            row) else None for i in range(len(headers))}
        rows.append(item)
    return success(import_questions_from_excel(db, rows, current_user.id))
